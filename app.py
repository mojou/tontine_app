# ============================================================
# APPLICATION DE GESTION DE TONTINE - VERSION CORRIGÉE
# ============================================================

import os
from datetime import datetime, timedelta, date
from decimal import Decimal
import random
import io

from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, abort, make_response, send_file
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from werkzeug.utils import secure_filename
from flask_wtf.csrf import CSRFProtect, generate_csrf
from flask_migrate import Migrate
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Importation des modules personnalisés
from config import Config
from forms import (
    LoginForm, MemberForm, TransactionForm, LoanRequestForm, LoanRepaymentForm,
    SanctionForm, AnnouncementForm, ProfileEditForm, TontineCycleForm, MemberEditForm,
    AideForm, AideApprovalForm, MeetingAttendanceForm, TontineBenefitForm, ReportForm,
    MemberFilterForm, TransactionFilterForm, PresenceTransactionForm, TontineTransactionForm,
    LoanApprovalForm, MemberRegistrationForm
)
from decorators import role_required, president_required, member_owner_required
from permissions import get_user_menu, has_permission
from utils import save_uploaded_file, log_activity, calculate_loan_interest, calculate_early_repayment_penalty, get_pagination_data

# Configuration des chemins
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# ============================================================
# INITIALISATION DE L'APPLICATION
# ============================================================

app = Flask(__name__,
            template_folder=os.path.join(BASE_DIR, 'app', 'templates'),
            static_folder=os.path.join(BASE_DIR, 'app', 'static'))

app.config.from_object(Config)

# Initialisation des extensions
csrf = CSRFProtect()
csrf.init_app(app)

from models import db
db.init_app(app)

migrate = Migrate(app, db)

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Veuillez vous connecter pour accéder à cette page.'

# ============================================================
# CONTEXT PROCESSOR
# ============================================================
# Dans app.py, après la création de l'app et avant les routes

@app.context_processor
def utility_processor():
    # Importation des fonctions depuis permissions.py
    from permissions import get_user_menu, has_permission, is_admin, is_bureau_member
    
    def get_current_year():
        return datetime.now().year

    def get_current_date_str():
        return datetime.now().strftime('%Y-%m-%d')

    return dict(
        # Fonctions de permissions
        get_user_menu=get_user_menu,
        has_permission=has_permission,
        is_admin=is_admin,
        is_bureau_member=is_bureau_member,
        
        # Utilitaires
        now=datetime.now(),
        current_year=get_current_year(),
        current_date=get_current_date_str(),
        csrf_token=generate_csrf
    )
# ============================================================
# CHARGEMENT DES MODÈLES
# ============================================================

from models import (
    User, Member, Transaction, Loan, TontineCycle, TontinePosition,
    Sanction, Announcement, AuditLog, AgendaItem, MeetingBeneficiary,
    Aide, TontineCycleDetail, CycleBeneficiary, MeetingAttendanceDetail,
    ContributionPlanning, CycleReport, CaisseBalance
)

# ============================================================
# USER LOADER
# ============================================================

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

# ============================================================
# FONCTIONS UTILITAIRES
# ============================================================

def set_member_choices(form):
    """Configure les choix de membres pour les formulaires"""
    members = Member.query.filter_by(is_active=True, status='ACTIF').all()
    choices = [(0, '-- Aucun --')] + [(m.id, m.full_name) for m in members]
    
    if hasattr(form, 'beneficiary_member_id'):
        form.beneficiary_member_id.choices = choices
    if hasattr(form, 'loan_member_id'):
        form.loan_member_id.choices = choices
    if hasattr(form, 'aid_member_id'):
        form.aid_member_id.choices = choices
    if hasattr(form, 'sanction_member_id'):
        form.sanction_member_id.choices = choices
    
    return members

def generate_cycle_contributions(cycle_id, ordered_member_ids, amount_per_member):
    """Génère automatiquement les échéances planifiées pour le cycle de tontine"""
    cycle = TontineCycleDetail.query.get(cycle_id)
    if not cycle:
        return False

    base_date = cycle.start_date
    
    for index, member_id in enumerate(ordered_member_ids):
        days_to_add = index * 14
        expected_draw_date = base_date + timedelta(days=days_to_add)
        
        contribution = ContributionPlanning(
            member_id=member_id,
            year=expected_draw_date.year,
            month=expected_draw_date.month,
            fortnight_number=1 if expected_draw_date.day <= 15 else 2,
            contribution_type='TONTINE',
            expected_amount=amount_per_member,
            expected_date=expected_draw_date,
            is_paid=False
        )
        db.session.add(contribution)
    
    db.session.commit()
    return True

# ============================================================
# CRÉATION DES TABLES ET DONNÉES INITIALES
# ============================================================

with app.app_context():
    upload_folder = app.config.get('UPLOAD_FOLDER', os.path.join(BASE_DIR, 'app', 'static', 'images'))
    os.makedirs(upload_folder, exist_ok=True)
    os.makedirs(os.path.join(BASE_DIR, 'reports'), exist_ok=True)
    os.makedirs(os.path.join(BASE_DIR, 'logs'), exist_ok=True)
    os.makedirs(os.path.join(BASE_DIR, 'instance'), exist_ok=True)
    
    db.create_all()
    
    default_users = [
        {'username': 'admin', 'password': 'admin123', 'role': 'SECRETAIRE', 'email': 'secretaire@tontine.com', 'first_name': 'Admin', 'last_name': 'Système', 'phone': '000000000'},
        {'username': 'president', 'password': 'president123', 'role': 'PRESIDENT', 'email': 'president@tontine.com', 'first_name': 'Président', 'last_name': 'Tontine', 'phone': '000000001'},
        {'username': 'tresorier', 'password': 'tresorier123', 'role': 'TRESORIER', 'email': 'tresorier@tontine.com', 'first_name': 'Trésorier', 'last_name': 'Tontine', 'phone': '000000002'},
        {'username': 'censeur', 'password': 'censeur123', 'role': 'CENSEUR', 'email': 'censeur@tontine.com', 'first_name': 'Censeur', 'last_name': 'Tontine', 'phone': '000000003'},
        {'username': 'communicateur', 'password': 'communicateur123', 'role': 'COMMUNICATION', 'email': 'communicateur@tontine.com', 'first_name': 'Communication', 'last_name': 'Tontine', 'phone': '000000004'},
    ]
    
    for user_data in default_users:
        if not User.query.filter_by(username=user_data['username']).first():
            member = Member(
                first_name=user_data['first_name'],
                last_name=user_data['last_name'],
                email=user_data['email'],
                phone=user_data['phone'],
                registration_date=date.today(),
                status='ACTIF',
                is_active=True,
                tontine_status='VERT',
                consecutive_failures=0,
                credit_balance=Decimal('0.00'),
                debit_balance=Decimal('0.00'),
                amount_to_receive=Decimal('0.00')
            )
            db.session.add(member)
            db.session.flush()
            
            user = User(
                username=user_data['username'],
                email=user_data['email'],
                role=user_data['role'],
                member_id=member.id,
                is_active=True
            )
            user.set_password(user_data['password'])
            db.session.add(user)
    
    db.session.commit()

# ============================================================
# ROUTES PRINCIPALES
# ============================================================
# ============================================================
# AUTHENTIFICATION ET DASHBOARD
# ============================================================

@app.route('/register', methods=['GET', 'POST'])
def register():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    
    form = MemberRegistrationForm()
    
    if form.validate_on_submit():
        existing_member = Member.query.filter_by(email=form.email.data).first()
        if existing_member:
            flash('Cet email est déjà utilisé. Veuillez vous connecter.', 'danger')
            return redirect(url_for('login'))
        
        existing_user = User.query.filter_by(username=form.username.data).first()
        if existing_user:
            flash('Ce nom d\'utilisateur est déjà pris.', 'danger')
            return render_template('register.html', form=form)
        
        member = Member(
            first_name=form.first_name.data,
            last_name=form.last_name.data,
            email=form.email.data,
            phone=form.phone.data,
            address=form.address.data,
            profession=form.profession.data,
            city=form.city.data,
            registration_date=date.today(),
            status='PENDING',
            is_active=False,
            tontine_status='VERT',
            consecutive_failures=0,
            credit_balance=Decimal('0.00'),
            debit_balance=Decimal('0.00'),
            amount_to_receive=Decimal('0.00'),
            group_type=None,
            position_in_group=None,
            has_received_benefit=False,
            chosen_tontine_amount=Decimal(str(form.tontine_amount.data)) if form.cotisation_type.data == 'TONTINE' and form.tontine_amount.data else Decimal('0.00')
        )
        
        db.session.add(member)
        db.session.flush()
        
        user = User(
            username=form.username.data,
            email=form.email.data,
            role='MEMBRE',
            member_id=member.id,
            is_active=False
        )
        user.set_password(form.password.data)
        db.session.add(user)
        db.session.commit()
        
        log_activity(None, 'PUBLIC', f"Nouvelle inscription: {member.full_name}", request.remote_addr)
        
        flash('Inscription réussie ! Votre dossier est en attente de validation.', 'success')
        return redirect(url_for('login'))
    
    return render_template('register.html', form=form)

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    
    form = LoginForm()
    if form.validate_on_submit():
        user = User.query.filter_by(username=form.username.data).first()
        if user and user.check_password(form.password.data) and user.is_active:
            login_user(user, remember=form.remember.data)
            user.last_login = datetime.utcnow()
            db.session.commit()
            
            flash(f'Bienvenue {user.member.full_name} ({user.get_role_display()}) !', 'success')
            next_page = request.args.get('next')
            return redirect(next_page) if next_page else redirect(url_for('dashboard'))
        else:
            flash('Nom d\'utilisateur ou mot de passe incorrect.', 'danger')
    
    return render_template('login.html', form=form)

@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('Vous avez été déconnecté.', 'info')
    return redirect(url_for('index'))

@app.route('/dashboard')
@login_required
def dashboard():
    member = current_user.member
    is_bureau = current_user.is_admin()
    
    if is_bureau:
        total_members = Member.query.filter_by(is_active=True).count()
        active_members = Member.query.filter_by(status='ACTIF', is_active=True).count()
        
        entrees = db.session.query(db.func.sum(Transaction.amount)).filter(
            Transaction.type.in_(['TONTINE', 'PRESENCE', 'SANCTION', 'REMBOURSEMENT', 'FONDS_CAISSE'])
        ).scalar() or 0
        
        sorties = db.session.query(db.func.sum(Transaction.amount)).filter(
            Transaction.type.in_(['SORTIE_LOAN', 'AIDES'])
        ).scalar() or 0
        
        loan_stats = db.session.query(
            db.func.count(Loan.id), 
            db.func.sum(Loan.total_amount - Loan.amount_paid)
        ).filter(Loan.status == 'ACTIF').first()
        
        total_active_loans = loan_stats[0] if loan_stats[0] else 0
        total_loan_amount = loan_stats[1] if loan_stats[1] else 0
        
        sanction_stats = db.session.query(db.func.count(Sanction.id), db.func.sum(Sanction.amount)).filter(
            Sanction.status == 'PENDING'
        ).first()
        pending_sanctions, total_sanctions_amount = sanction_stats if sanction_stats[0] else (0, 0)
        
        next_cycle = TontineCycleDetail.query.filter_by(status='EN_COURS').first()
        next_beneficiary = next_cycle.get_next_beneficiary() if next_cycle else None
        
        stats = {
            'total_members': total_members,
            'active_members': active_members,
            'total_cotisations': float(entrees),
            'fonds_caisse': float(entrees - sorties),
            'total_active_loans': int(total_active_loans or 0),
            'total_loan_amount': float(total_loan_amount or 0),
            'pending_sanctions': int(pending_sanctions or 0),
            'total_sanctions_amount': float(total_sanctions_amount or 0),
            'next_beneficiary': next_beneficiary.full_name if next_beneficiary else None,
            'upcoming_meetings': Announcement.query.filter(
                Announcement.announcement_type == 'REUNION',
                Announcement.event_date >= date.today()
            ).order_by(Announcement.event_date).limit(5).all()
        }
    else:
        my_loans_total = db.session.query(db.func.sum(Loan.total_amount - Loan.amount_paid)).filter(
            Loan.member_id == member.id, 
            Loan.status == 'ACTIF'
        ).scalar() or 0
        
        stats = {
            'my_cotisations': float(member.total_savings or 0),
            'my_presence': float(member.total_presence_paid or 0),
            'my_loans': float(my_loans_total or 0),
            'my_sanctions': float(member.total_sanctions_pending or 0),
            'my_position': f"Position {member.position_in_group}/{member.group_type}" if member.group_type else "Non affecté",
            'my_status': member.tontine_status,
            'my_credit': float(member.credit_balance or 0),
            'my_debit': float(member.debit_balance or 0)
        }
    
    query = Transaction.query if is_bureau else Transaction.query.filter_by(member_id=member.id)
    recent_transactions = query.order_by(Transaction.date.desc()).limit(10).all()
    
    return render_template('dashboard.html', 
                           stats=stats, 
                           is_bureau=is_bureau,
                           recent_transactions=recent_transactions,
                           member=member)


# ============================================================
#index
@app.route('/')
def index():
    """Page d'accueil publique"""
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    
    # Statistiques pour la page d'accueil
    total_members = Member.query.filter_by(is_active=True).count()
    active_loans = Loan.query.filter_by(status='ACTIF').count()
    total_collected = db.session.query(db.func.sum(Transaction.amount)).filter(
        Transaction.type.in_(['TONTINE', 'PRESENCE'])
    ).scalar() or 0
    
    stats = {
        'total_members': total_members,
        'active_loans': active_loans,
        'total_collected': float(total_collected),
    }
    
    return render_template('index.html', stats=stats)
# ============================================================
# ============================================================
# GESTION DES MEMBRES
# ============================================================

@app.route('/members')
@login_required
@role_required(['SECRETAIRE', 'PRESIDENT', 'TRESORIER', 'COMMUNICATION', 'CENSEUR'])
def members():
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 10, type=int)
    search = request.args.get('search', '')
    status_filter = request.args.get('status', '')
    tontine_status_filter = request.args.get('tontine_status', '')
    group_type_filter = request.args.get('group_type', '')
    show_inactive = request.args.get('show_inactive', 'false') == 'true'
    
    query = Member.query
    
    if not show_inactive:
        query = query.filter_by(is_active=True)
    
    if search:
        query = query.filter(
            db.or_(
                Member.first_name.ilike(f'%{search}%'),
                Member.last_name.ilike(f'%{search}%'),
                Member.phone.contains(search),
                Member.email.ilike(f'%{search}%')
            )
        )
    
    if status_filter:
        query = query.filter_by(status=status_filter)
    if tontine_status_filter:
        query = query.filter_by(tontine_status=tontine_status_filter)
    if group_type_filter:
        query = query.filter_by(group_type=group_type_filter)
    
    pagination = query.order_by(Member.last_name).paginate(page=page, per_page=per_page, error_out=False)
    pagination_data = get_pagination_data(pagination)
    
    form = MemberForm()
    
    return render_template('members.html', 
                           members=pagination_data['items'], 
                           pagination=pagination_data, 
                           search=search,
                           status_filter=status_filter,
                           tontine_status_filter=tontine_status_filter,
                           group_type_filter=group_type_filter,
                           show_inactive=show_inactive,
                           form=form)

@app.route('/members/<int:member_id>')
@login_required
@member_owner_required
def member_detail(member_id):
    member = Member.query.get_or_404(member_id)
    page = request.args.get('page', 1, type=int)
    per_page = 10
    
    transactions_query = Transaction.query.filter_by(member_id=member_id).order_by(Transaction.date.desc())
    transactions_pagination = transactions_query.paginate(page=page, per_page=per_page, error_out=False)
    transactions_pagination_data = get_pagination_data(transactions_pagination)
    
    loans = Loan.query.filter_by(member_id=member_id).all()
    sanctions = Sanction.query.filter_by(member_id=member_id, status='PENDING').all()
    aides = Aide.query.filter_by(member_id=member_id).all()
    cycle_benefits = CycleBeneficiary.query.filter_by(member_id=member_id).all()
    
    total_contributions = db.session.query(db.func.sum(Transaction.amount)).filter(
        Transaction.member_id == member_id,
        Transaction.type.in_(['TONTINE', 'PRESENCE', 'FONDS_CAISSE'])
    ).scalar() or 0
    
    total_loans = db.session.query(db.func.sum(Loan.amount)).filter(
        Loan.member_id == member_id,
        Loan.status == 'ACTIF'
    ).scalar() or 0
    
    can_edit = current_user.role == 'PRESIDENT'
    can_change_role = current_user.role == 'PRESIDENT'
    
    return render_template('member_detail.html',
                           member=member,
                           transactions=transactions_pagination_data['items'],
                           transactions_pagination=transactions_pagination_data,
                           loans=loans,
                           sanctions=sanctions,
                           aides=aides,
                           cycle_benefits=cycle_benefits,
                           total_contributions=total_contributions,
                           total_loans=total_loans,
                           can_edit=can_edit,
                           can_change_role=can_change_role)

@app.route('/members/add', methods=['GET', 'POST'])
@login_required
@role_required(['PRESIDENT', 'SECRETAIRE'])
def add_member():
    form = MemberForm()
    
    # UTILISEZ LA VALIDATION DU FORMULAIRE !!!
    if form.validate_on_submit():
        # Les données sont automatiquement validées et nettoyées
        first_name = form.first_name.data
        last_name = form.last_name.data
        email = form.email.data
        phone = form.phone.data
        address = form.address.data
        profession = form.profession.data
        city = form.city.data
        group_type = form.group_type.data
        position_in_group = form.position_in_group.data
        chosen_tontine_amount = form.chosen_tontine_amount.data
        username = form.username.data
        role = form.role.data
        password = form.password.data
        confirm_password = form.confirm_password.data
        
        # Vérification supplémentaire des mots de passe
        if password != confirm_password:
            flash('Les mots de passe ne correspondent pas.', 'danger')
            return render_template('member_form.html', title="Ajouter un membre", form=form)
        
        # Vérification des doublons
        existing_member = Member.query.filter_by(email=email).first()
        if existing_member:
            flash(f'L\'email "{email}" est déjà utilisé.', 'danger')
            return render_template('member_form.html', title="Ajouter un membre", form=form)
        
        existing_user = User.query.filter_by(username=username).first()
        if existing_user:
            flash(f'Le nom d\'utilisateur "{username}" est déjà pris.', 'danger')
            return render_template('member_form.html', title="Ajouter un membre", form=form)
        
        # Création du membre
        member = Member(
            first_name=first_name, last_name=last_name, email=email,
            phone=phone, address=address, profession=profession, city=city,
            registration_date=date.today(), status='ACTIF', is_active=True,
            tontine_status='VERT', consecutive_failures=0,
            credit_balance=Decimal('0.00'), debit_balance=Decimal('0.00'),
            amount_to_receive=Decimal('0.00'), group_type=group_type,
            position_in_group=int(position_in_group) if position_in_group else None,
            has_received_benefit=False,
            chosen_tontine_amount=Decimal(str(chosen_tontine_amount)) if chosen_tontine_amount else None
        )
        
        if 'photo' in request.files:
            photo = request.files['photo']
            if photo and photo.filename:
                filename = save_uploaded_file(photo, app.config['UPLOAD_FOLDER'])
                if filename:
                    member.photo = filename
        
        db.session.add(member)
        db.session.flush()
        
        user = User(username=username, email=email, role=role,
                    member_id=member.id, is_active=True)
        user.set_password(password)
        db.session.add(user)
        db.session.commit()
        
        log_activity(current_user.id, current_user.role, f"Ajout membre: {member.full_name}", request.remote_addr)
        flash('Membre ajouté avec succès !', 'success')
        return redirect(url_for('members'))
    
    # Si le formulaire n'est pas valide, afficher les erreurs
    for field, errors in form.errors.items():
        for error in errors:
            flash(f'Erreur dans le champ {getattr(form, field).label.text}: {error}', 'danger')
    
    return render_template('member_form.html', title="Ajouter un membre", form=form)


@app.route('/members/<int:member_id>/edit', methods=['GET', 'POST'])
@login_required
@president_required
def edit_member(member_id):
    member = Member.query.get_or_404(member_id)
    form = MemberEditForm(obj=member)
    
    if form.validate_on_submit():
        member.first_name = form.first_name.data
        member.last_name = form.last_name.data
        member.email = form.email.data
        member.phone = form.phone.data
        member.address = form.address.data
        member.profession = form.profession.data
        member.city = form.city.data
        member.status = form.status.data
        
        if form.photo.data:
            filename = save_uploaded_file(form.photo.data, app.config['UPLOAD_FOLDER'])
            if filename:
                member.photo = filename
        
        db.session.commit()
        log_activity(current_user.id, current_user.role, f"Modification membre: {member.full_name}", request.remote_addr)
        flash('Membre modifié avec succès !', 'success')
        return redirect(url_for('member_detail', member_id=member.id))
    
    return render_template('member_form.html', form=form, member=member, title="Modifier membre")

@app.route('/members/<int:member_id>/delete', methods=['POST'])
@login_required
@president_required
def delete_member(member_id):
    member = Member.query.get_or_404(member_id)
    member.is_active = False
    
    user = User.query.filter_by(member_id=member.id).first()
    if user:
        user.is_active = False
    
    db.session.commit()
    log_activity(current_user.id, current_user.role, f"Désactivation membre: {member.full_name}", request.remote_addr)
    flash('Membre désactivé avec succès !', 'success')
    return redirect(url_for('members'))

@app.route('/members/<int:member_id>/activate', methods=['POST'])
@login_required
@role_required(['TRESORIER', 'PRESIDENT'])
def activate_member(member_id):
    member = Member.query.get_or_404(member_id)
    
    if member.status == 'PENDING':
        member.status = 'ACTIF'
        member.is_active = True
        
        user = User.query.filter_by(member_id=member.id).first()
        if user:
            user.is_active = True
        
        db.session.commit()
        log_activity(current_user.id, current_user.role, f"Activation membre: {member.full_name}", request.remote_addr)
        flash(f'Membre {member.full_name} activé avec succès !', 'success')
    else:
        flash('Ce membre n\'est pas en attente de validation.', 'warning')
    
    return redirect(url_for('members'))


# ============================================================
# GESTION DES MEMBRES
# ============================================================

# ... vos routes existantes ...

@app.route('/member/<int:member_id>/update_status', methods=['POST'])
@login_required
def update_status(member_id):
    """Mettre à jour le statut tontine d'un membre (VERT/ORANGE/ROUGE)"""
    if current_user.role not in ['SECRETAIRE', 'PRESIDENT', 'TRESORIER']:
        flash('Vous n\'avez pas les droits nécessaires.', 'danger')
        return redirect(url_for('member_detail', member_id=member_id))
    
    member = Member.query.get_or_404(member_id)
    new_status = request.form.get('tontine_status')
    
    if new_status and new_status in ['VERT', 'ORANGE', 'ROUGE']:
        member.tontine_status = new_status
        db.session.commit()
        log_activity(current_user.id, current_user.role, f"Statut tontine de {member.full_name} -> {new_status}", request.remote_addr)
        flash(f'Statut tontine de {member.full_name} mis à jour : {new_status}', 'success')
    else:
        flash('Statut invalide.', 'danger')
    
    return redirect(url_for('member_detail', member_id=member_id))


@app.route('/members/<int:member_id>/update_member_status', methods=['POST'])
@login_required
@role_required(['PRESIDENT', 'SECRETAIRE'])
def update_member_status(member_id):
    """Mettre à jour le statut général du membre (ACTIF/SUSPENDU/EXCLU)"""
    member = Member.query.get_or_404(member_id)
    new_status = request.form.get('status')
    valid_statuses = ['ACTIF', 'SUSPENDU', 'EXCLU']
    
    if new_status and new_status in valid_statuses:
        member.status = new_status
        db.session.commit()
        log_activity(current_user.id, current_user.role, f"Statut de {member.full_name} -> {new_status}", request.remote_addr)
        
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': True, 'status': new_status})
        
        flash(f'Statut de {member.full_name} mis à jour avec succès.', 'success')
    else:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': False, 'message': 'Statut invalide'})
        flash('Statut invalide.', 'danger')
    
    return redirect(url_for('member_detail', member_id=member_id))


@app.route('/members/<int:member_id>/change-role', methods=['POST'])
@login_required
@president_required
def change_member_role(member_id):
    """Changer le rôle d'un membre (MEMBRE/PRESIDENT/SECRETAIRE/TRESORIER/CENSEUR/COMMUNICATION)"""
    member = Member.query.get_or_404(member_id)
    new_role = request.form.get('role')
    valid_roles = ['MEMBRE', 'PRESIDENT', 'SECRETAIRE', 'TRESORIER', 'CENSEUR', 'COMMUNICATION']
    
    if new_role not in valid_roles:
        flash('Rôle invalide.', 'danger')
        return redirect(url_for('member_detail', member_id=member_id))
    
    user = User.query.filter_by(member_id=member.id).first()
    if user:
        old_role = user.role
        user.role = new_role
        db.session.commit()
        log_activity(current_user.id, current_user.role, f"Changement rôle de {member.full_name}: {old_role} -> {new_role}", request.remote_addr)
        flash(f'Rôle de {member.full_name} modifié en {new_role}.', 'success')
    else:
        flash('Ce membre n\'a pas de compte utilisateur.', 'danger')
    
    return redirect(url_for('member_detail', member_id=member_id))


@app.route('/members/<int:member_id>/reactivate', methods=['POST'])
@login_required
@president_required
def reactivate_member(member_id):
    """Réactiver un membre désactivé"""
    member = Member.query.get_or_404(member_id)
    member.is_active = True
    
    user = User.query.filter_by(member_id=member.id).first()
    if user:
        user.is_active = True
    
    db.session.commit()
    log_activity(current_user.id, current_user.role, f"Réactivation membre: {member.full_name}", request.remote_addr)
    flash('Membre réactivé avec succès !', 'success')
    return redirect(url_for('members'))
# ============================================================
# GESTION DES TRANSACTIONS
# ============================================================

@app.route('/transactions')
@login_required
def transactions():
    if not current_user.is_admin() and request.args.get('member_id', type=int) != current_user.member_id:
        return redirect(url_for('transactions', member_id=current_user.member_id))

    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 10, type=int)
    type_filter = request.args.get('type', '')
    date_debut = request.args.get('date_debut', '')
    date_fin = request.args.get('date_fin', '')
    member_id = request.args.get('member_id', type=int)
    
    query = Transaction.query
    
    if type_filter:
        query = query.filter_by(type=type_filter)
    if date_debut:
        query = query.filter(Transaction.date >= datetime.strptime(date_debut, '%Y-%m-%d').date())
    if date_fin:
        query = query.filter(Transaction.date <= datetime.strptime(date_fin, '%Y-%m-%d').date())
    if member_id:
        query = query.filter_by(member_id=member_id)
    
    pagination = query.order_by(Transaction.date.desc()).paginate(page=page, per_page=per_page, error_out=False)
    pagination_data = get_pagination_data(pagination)
    
    summary = {}
    for t_type in ['TONTINE', 'PRESENCE', 'SANCTION', 'REMBOURSEMENT', 'AIDE', 'FONDS_CAISSE', 'SORTIE_LOAN', 'BENEFICE_TONTINE']:
        summary[t_type] = db.session.query(db.func.sum(Transaction.amount)).filter_by(type=t_type).scalar() or 0
    
    members_list = Member.query.filter_by(is_active=True).all() if current_user.is_admin() else []
    
    return render_template('transactions.html',
                           transactions=pagination_data['items'],
                           pagination=pagination_data,
                           summary=summary,
                           type_filter=type_filter,
                           members=members_list,
                           date_debut=date_debut,
                           date_fin=date_fin,
                           selected_member=member_id)

@app.route('/transactions/add', methods=['GET', 'POST'])
@login_required
@role_required(['SECRETAIRE', 'PRESIDENT', 'TRESORIER'])
def add_transaction():
    form = TransactionForm()
    form.member_id.choices = [(0, 'Choisir un membre')] + [(m.id, m.full_name) for m in Member.query.filter_by(is_active=True).order_by(Member.last_name).all()]
    
    if form.validate_on_submit():
        if form.member_id.data == 0:
            flash('Veuillez sélectionner un membre valide.', 'danger')
            return render_template('transaction_form.html', form=form, title="Ajouter une transaction")

        transaction = Transaction(
            member_id=form.member_id.data,
            type=form.type.data,
            amount=Decimal(str(form.amount.data)),
            description=form.description.data,
            date=date.today(),
            created_by=current_user.id,
            payment_mode=form.payment_mode.data if hasattr(form, 'payment_mode') else 'ESPECE'
        )
        db.session.add(transaction)

        if form.type.data == 'TONTINE':
            member = Member.query.get(form.member_id.data)
            if member and not member.has_paid_fonds_caisse:
                fonds_caisse = Transaction(
                    member_id=member.id,
                    type='FONDS_CAISSE',
                    amount=Decimal('5000.00'),
                    description="Fonds de Caisse obligatoire lié à la Tontine",
                    date=date.today(),
                    created_by=current_user.id,
                    payment_mode='ESPECE'
                )
                db.session.add(fonds_caisse)
                flash('Fonds de caisse obligatoire enregistré !', 'info')
        
        db.session.commit()
        log_activity(current_user.id, current_user.role, f"Ajout transaction: {form.type.data}", request.remote_addr)
        flash('Transaction enregistrée avec succès !', 'success')
        return redirect(url_for('transactions'))
    
    return render_template('transaction_form.html', form=form, title="Ajouter une transaction")


@app.route('/transactions/<int:transaction_id>/edit', methods=['GET', 'POST'])
@login_required
@role_required(['TRESORIER', 'PRESIDENT'])
def edit_transaction(transaction_id):
    """Modifier une transaction existante"""
    transaction = Transaction.query.get_or_404(transaction_id)
    form = TransactionForm(obj=transaction)
    
    # Remplir les choix du select
    form.member_id.choices = [(0, 'Choisir un membre')] + [(m.id, m.full_name) for m in Member.query.filter_by(is_active=True).order_by(Member.last_name).all()]
    
    if form.validate_on_submit():
        try:
            transaction.member_id = form.member_id.data
            transaction.type = form.type.data
            transaction.amount = Decimal(str(form.amount.data))
            transaction.description = form.description.data
            transaction.payment_mode = form.payment_mode.data if hasattr(form, 'payment_mode') else transaction.payment_mode
            
            db.session.commit()
            
            log_activity(current_user.id, current_user.role, f"Modification transaction #{transaction_id}", request.remote_addr)
            flash('Transaction modifiée avec succès !', 'success')
            return redirect(url_for('transactions'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Erreur lors de la modification : {str(e)}', 'danger')
    
    return render_template('transaction_form.html', form=form, transaction=transaction, title="Modifier une transaction")


@app.route('/transactions/<int:transaction_id>/delete', methods=['POST'])
@login_required
@role_required(['PRESIDENT'])
def delete_transaction(transaction_id):
    """Supprimer une transaction"""
    transaction = Transaction.query.get_or_404(transaction_id)
    
    try:
        db.session.delete(transaction)
        db.session.commit()
        
        log_activity(current_user.id, current_user.role, f"Suppression transaction #{transaction_id}", request.remote_addr)
        flash('Transaction supprimée avec succès !', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Erreur lors de la suppression : {str(e)}', 'danger')
    
    return redirect(url_for('transactions'))


@app.route('/transactions/<int:transaction_id>/view')
@login_required
def view_transaction(transaction_id):
    """Voir le détail d'une transaction"""
    transaction = Transaction.query.get_or_404(transaction_id)
    
    # Vérifier les droits d'accès
    if not current_user.is_admin() and transaction.member_id != current_user.member_id:
        flash('Vous n\'avez pas accès à cette transaction.', 'danger')
        return redirect(url_for('transactions'))
    
    return render_template('transaction_detail.html', transaction=transaction)

# ============================================================
# GESTION DES EMPRUNTS
# ============================================================

@app.route('/loans')
@login_required
def loans():
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 10, type=int)
    status_filter = request.args.get('status', '')
    overdue_filter = request.args.get('overdue', '')
    search_query = request.args.get('search', '')
    sort_column = request.args.get('sort', 'id')
    sort_order = request.args.get('order', 'desc')
    
    query = Loan.query
    
    if not current_user.is_admin():
        query = query.filter_by(member_id=current_user.member_id)
    
    if status_filter:
        query = query.filter(Loan.status == status_filter)
        
    if overdue_filter == 'yes':
        query = query.filter(Loan.status.in_(['ACTIF', 'OVERDUE']), Loan.end_date < date.today())
    elif overdue_filter == 'no':
        query = query.filter(db.or_(
            Loan.status == 'REMBOURSE',
            db.and_(Loan.status == 'ACTIF', Loan.end_date >= date.today())
        ))
        
    if search_query and current_user.is_admin():
        query = query.join(Member).filter(
            db.or_(
                Member.first_name.ilike(f'%{search_query}%'), 
                Member.last_name.ilike(f'%{search_query}%')
            )
        )
        
    if sort_column == 'member_name':
        if not search_query:
            query = query.join(Member)
        if sort_order == 'asc':
            query = query.order_by(Member.first_name.asc(), Member.last_name.asc())
        else:
            query = query.order_by(Member.first_name.desc(), Member.last_name.desc())
    elif sort_column == 'amount':
        query = query.order_by(Loan.amount.asc() if sort_order == 'asc' else Loan.amount.desc())
    else:
        query = query.order_by(Loan.id.desc())
        
    pagination = query.paginate(page=page, per_page=per_page, error_out=False)
    
    for loan in pagination.items:
        if loan.is_overdue and loan.status == 'ACTIF':
            loan.status = 'OVERDUE'
    
    db.session.commit()
    
    members_list = []
    if current_user.is_admin():
        members_list = Member.query.filter_by(is_active=True).all()
    else:
        if current_user.member:
            members_list = [current_user.member]

    return render_template(
        'loans.html', 
        loans=pagination.items, 
        pagination=get_pagination_data(pagination), 
        members=members_list,
        current_sort=sort_column,
        current_order=sort_order
    )

@app.route('/loan/request', methods=['POST'])
@login_required
def loan_request():
    """Soumettre une demande d'emprunt"""
    member_id = request.form.get('member_id', type=int)
    amount_str = request.form.get('amount')
    duration_months = request.form.get('duration_months', type=int, default=3)
    interest_rate_str = request.form.get('interest_rate', '5.0')
    purpose = request.form.get('purpose', 'Emprunt standard')
    
    # Vérifier les droits
    if not current_user.is_admin() and current_user.member_id != member_id:
        flash("Action non autorisée.", "danger")
        return redirect(url_for('loans'))
    
    member = Member.query.get_or_404(member_id)
    amount = Decimal(amount_str) if amount_str else Decimal('0.00')
    interest_rate = Decimal(interest_rate_str) if interest_rate_str else Decimal('5.00')
    
    # Vérifier l'éligibilité
    if hasattr(member, 'is_eligible_for_loan'):
        is_eligible, reason = member.is_eligible_for_loan(amount)
        if not is_eligible:
            flash(f"Échec d'éligibilité : {reason}", "danger")
            return redirect(url_for('loans'))
    
    # Calculer les intérêts
    interest = amount * (interest_rate / Decimal('100'))
    total_amount = amount + interest
    
    # Créer la demande d'emprunt
    loan = Loan(
        member_id=member.id,
        amount=amount,
        interest=interest,
        total_amount=total_amount,
        amount_paid=Decimal('0.00'),
        request_date=date.today(),
        approval_date=None,
        start_date=None,
        end_date=date.today() + timedelta(days=duration_months * 30),
        status='PENDING',
        description=purpose,
        approved_by=None,
        created_by=current_user.id,
        created_at=datetime.utcnow()
    )
    
    db.session.add(loan)
    db.session.commit()
    
    log_activity(current_user.id, current_user.role, f"Demande d'emprunt de {amount:,.0f} FCFA", request.remote_addr)
    flash(f"Demande d'emprunt de {amount:,.0f} FCFA soumise avec succès.", "success")
    return redirect(url_for('loans'))


@app.route('/loans/<int:loan_id>/approve', methods=['POST'])
@login_required
def approve_loan(loan_id):
    """Approuver une demande d'emprunt"""
    if getattr(current_user, 'role', None) not in ['PRESIDENT', 'TRESORIER']:
        abort(403)
    
    loan = Loan.query.get_or_404(loan_id)
    
    if loan.status != 'PENDING':
        flash("Cet emprunt n'est pas en attente.", "warning")
        return redirect(url_for('loans'))
    
    loan.status = 'ACTIF'
    loan.approval_date = date.today()
    loan.approved_by = current_user.id
    loan.start_date = date.today()
    
    # Créer la transaction de décaissement
    transaction = Transaction(
        member_id=loan.member_id,
        type='SORTIE_LOAN',
        amount=loan.amount,
        description=f"Décaissement emprunt #{loan.id}",
        date=date.today(),
        created_by=current_user.id,
        payment_mode='ESPECE'
    )
    db.session.add(transaction)
    db.session.commit()
    
    log_activity(current_user.id, current_user.role, f"Emprunt #{loan.id} approuvé", request.remote_addr)
    flash('Le prêt a été approuvé avec succès.', 'success')
    return redirect(url_for('loans'))


@app.route('/loans/<int:loan_id>/reject', methods=['POST'])
@login_required
def reject_loan(loan_id):
    """Rejeter une demande d'emprunt"""
    if getattr(current_user, 'role', None) not in ['PRESIDENT', 'TRESORIER']:
        abort(403)
    
    loan = Loan.query.get_or_404(loan_id)
    
    if loan.status != 'PENDING':
        flash("Impossible de rejeter un emprunt déjà traité.", "warning")
        return redirect(url_for('loans'))
    
    rejection_reason = request.form.get('rejection_reason', 'Refusé par le bureau')
    loan.status = 'REJECTED'
    loan.description = rejection_reason
    db.session.commit()
    
    log_activity(current_user.id, current_user.role, f"Emprunt #{loan.id} rejeté", request.remote_addr)
    flash('La demande d\'emprunt a été rejetée.', 'warning')
    return redirect(url_for('loans'))


@app.route('/loans/<int:loan_id>/repay', methods=['GET', 'POST'])
@login_required
def repay_loan(loan_id):
    """Rembourser un emprunt"""
    if getattr(current_user, 'role', None) not in ['SECRETAIRE', 'PRESIDENT', 'TRESORIER']:
        abort(403)
    
    loan = Loan.query.get_or_404(loan_id)
    
    if loan.status not in ['ACTIF', 'OVERDUE']:
        flash('Ce prêt n\'est pas actif ou est déjà soldé.', 'warning')
        return redirect(url_for('loans'))
    
    if request.method == 'GET':
        return render_template('loan_repay_form.html', loan=loan, title="Rembourser un emprunt")
    
    # Traitement POST
    amount_paid_str = request.form.get('amount')
    if not amount_paid_str or Decimal(amount_paid_str) <= 0:
        flash('Veuillez entrer un montant de remboursement valide.', 'danger')
        return redirect(url_for('loans'))
    
    payment_amount = Decimal(amount_paid_str)
    remaining = loan.total_amount - loan.amount_paid
    
    if payment_amount > remaining:
        flash(f"Le montant saisi dépasse le solde restant dû ({remaining:,.0f} FCFA).", "danger")
        return redirect(url_for('loans'))
    
    loan.amount_paid += payment_amount
    
    if loan.amount_paid >= loan.total_amount:
        loan.status = 'REMBOURSE'
    
    transaction = Transaction(
        member_id=loan.member_id,
        type='REMBOURSEMENT',
        amount=payment_amount,
        description=f"Remboursement emprunt #{loan.id}",
        date=date.today(),
        created_by=current_user.id,
        payment_mode=request.form.get('payment_mode', 'ESPECE')
    )
    db.session.add(transaction)
    db.session.commit()
    
    log_activity(current_user.id, current_user.role, f"Remboursement emprunt #{loan.id}", request.remote_addr)
    flash("Le remboursement a été enregistré.", "success")
    return redirect(url_for('loans'))


# ============================================================
# GESTION DES SANCTIONS
# ============================================================

# ============================================================
# GESTION DES SANCTIONS (AVEC VALIDATION)
# ============================================================

@app.route('/sanctions', methods=['GET'])
@login_required
def sanctions():
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 10, type=int)
    status_filter = request.args.get('status', '')
    
    query = Sanction.query
    if not current_user.is_admin():
        query = query.filter_by(member_id=current_user.member_id)
    if status_filter:
        query = query.filter_by(status=status_filter)
        
    pagination = query.order_by(Sanction.created_at.desc()).paginate(page=page, per_page=per_page, error_out=False)
    pagination_data = get_pagination_data(pagination)
    
    members_list = Member.query.filter_by(is_active=True).order_by(Member.last_name).all() if current_user.is_admin() else []
    
    # Créer le formulaire pour l'ajout
    form = SanctionForm()
    form.member_id.choices = [(m.id, m.full_name) for m in Member.query.filter_by(is_active=True, status='ACTIF').order_by(Member.last_name).all()]
    
    return render_template('sanctions.html', 
                           sanctions=pagination_data['items'], 
                           pagination=pagination_data, 
                           members=members_list,
                           status_filter=status_filter,
                           form=form)


@app.route('/sanctions/add', methods=['GET', 'POST'])
@login_required
@role_required(['CENSEUR', 'PRESIDENT', 'SECRETAIRE'])
def add_sanction():
    """Ajouter une nouvelle sanction avec validation"""
    form = SanctionForm()
    form.member_id.choices = [(m.id, m.full_name) for m in Member.query.filter_by(is_active=True, status='ACTIF').order_by(Member.last_name).all()]
    
    # Afficher le formulaire en GET
    if request.method == 'GET':
        return render_template('sanction_form.html', form=form, title="Ajouter une sanction")
    
    # Traitement POST avec validation
    if form.validate_on_submit():
        try:
            sanction = Sanction(
                member_id=form.member_id.data,
                type_sanction=form.type_sanction.data,
                amount=Decimal(str(form.amount.data)),
                description=form.description.data,
                sanction_date=form.sanction_date.data,
                status='PENDING'
            )
            db.session.add(sanction)
            db.session.commit()
            
            member = Member.query.get(form.member_id.data)
            log_activity(current_user.id, current_user.role, f"Sanction infligée à {member.full_name} : {form.amount.data} FCFA", request.remote_addr)
            flash(f"Amende de {form.amount.data:,.0f} FCFA enregistrée avec succès pour {member.full_name}.", "success")
            return redirect(url_for('sanctions'))
            
        except Exception as e:
            db.session.rollback()
            flash(f"Erreur lors de l'enregistrement : {str(e)}", "danger")
    else:
        # Afficher les erreurs de validation
        for field, errors in form.errors.items():
            for error in errors:
                field_label = getattr(form, field).label.text if hasattr(form, field) else field
                flash(f'Erreur dans le champ "{field_label}": {error}', 'danger')
    
    return render_template('sanction_form.html', form=form, title="Ajouter une sanction")


@app.route('/sanctions/<int:sanction_id>/edit', methods=['GET', 'POST'])
@login_required
@role_required(['CENSEUR', 'PRESIDENT'])
def edit_sanction(sanction_id):
    """Modifier une sanction existante"""
    sanction = Sanction.query.get_or_404(sanction_id)
    
    if sanction.status == 'PAID':
        flash('Impossible de modifier une sanction déjà payée.', 'danger')
        return redirect(url_for('sanctions'))
    
    form = SanctionForm(obj=sanction)
    form.member_id.choices = [(m.id, m.full_name) for m in Member.query.filter_by(is_active=True).order_by(Member.last_name).all()]
    
    if form.validate_on_submit():
        try:
            sanction.member_id = form.member_id.data
            sanction.type_sanction = form.type_sanction.data
            sanction.amount = Decimal(str(form.amount.data))
            sanction.description = form.description.data
            sanction.sanction_date = form.sanction_date.data
            
            db.session.commit()
            
            log_activity(current_user.id, current_user.role, f"Modification sanction #{sanction_id}", request.remote_addr)
            flash('Sanction modifiée avec succès.', 'success')
            return redirect(url_for('sanctions'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Erreur : {str(e)}', 'danger')
    else:
        for field, errors in form.errors.items():
            for error in errors:
                flash(f'Erreur: {error}', 'danger')
    
    return render_template('sanction_form.html', form=form, sanction=sanction, title="Modifier la sanction")


@app.route('/sanctions/<int:sanction_id>/pay', methods=['POST'])
@login_required
@role_required(['TRESORIER', 'PRESIDENT'])
def pay_sanction(sanction_id):
    """Payer une sanction"""
    sanction = Sanction.query.get_or_404(sanction_id)
    
    if sanction.status == 'PAID':
        flash('Cette amende a déjà été payée.', 'warning')
        return redirect(url_for('sanctions'))
    
    # Validation du mode de paiement
    payment_mode = request.form.get('payment_mode', 'ESPECE')
    valid_modes = ['ESPECE', 'ORANGE_MONEY', 'MTN_MOBILE', 'VIREMENT']
    if payment_mode not in valid_modes:
        payment_mode = 'ESPECE'
    
    sanction.status = 'PAID'
    
    transaction = Transaction(
        member_id=sanction.member_id,
        type='SANCTION',
        amount=sanction.amount,
        description=f"Paiement sanction #{sanction.id} - {sanction.description[:100]}",
        date=date.today(),
        created_by=current_user.id,
        payment_mode=payment_mode
    )
    db.session.add(transaction)
    db.session.commit()
    
    log_activity(current_user.id, current_user.role, f"Paiement sanction #{sanction.id} - {sanction.amount} FCFA", request.remote_addr)
    flash(f"Le paiement de l'amende de {sanction.amount:,.0f} FCFA a été encaissé.", "success")
    return redirect(url_for('sanctions'))


@app.route('/sanctions/<int:sanction_id>/delete', methods=['POST'])
@login_required
@role_required(['CENSEUR', 'PRESIDENT'])
def delete_sanction(sanction_id):
    """Supprimer une sanction (uniquement si non payée)"""
    sanction = Sanction.query.get_or_404(sanction_id)
    
    if sanction.status == 'PAID':
        flash('Impossible de supprimer une sanction déjà payée.', 'danger')
        return redirect(url_for('sanctions'))
    
    try:
        db.session.delete(sanction)
        db.session.commit()
        
        log_activity(current_user.id, current_user.role, f"Suppression sanction #{sanction_id}", request.remote_addr)
        flash('Sanction supprimée avec succès.', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Erreur lors de la suppression : {str(e)}', 'danger')
    
    return redirect(url_for('sanctions'))
# ============================================================
# GESTION DES RÉUNIONS
# ============================================================

# ============================================================
# GESTION DES RÉUNIONS (CRUD COMPLET)
# ============================================================

# ============================================================
# GESTION DES RÉUNIONS (CRUD COMPLET AVEC VALIDATION)
# ============================================================

@app.template_filter('nl2br')
def nl2br_filter(text):
    """
    Convertit les sauts de ligne en balises HTML <br>
    Échappe d'abord le HTML pour éviter les injections XSS
    """
    if not text:
        return ''
    from markupsafe import escape
    # Échappe d'abord le HTML
    escaped = escape(text)
    # Remplace les sauts de ligne
    return escaped.replace('\n', '<br>\n')

@app.route('/meetings')
@login_required
def meetings():
    """Liste des réunions"""
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 10, type=int)
    
    pagination = Announcement.query.filter(
        Announcement.announcement_type == 'REUNION'
    ).order_by(Announcement.event_date.desc().nullslast()).paginate(
        page=page, per_page=per_page, error_out=False
    )
    
    return render_template('meetings.html', 
                           meetings=pagination.items, 
                           pagination=pagination)


@app.route('/meetings/<int:meeting_id>')
@login_required
def meeting_detail(meeting_id):
    """Détail d'une réunion"""
    meeting = Announcement.query.get_or_404(meeting_id)
    agenda_items = AgendaItem.query.filter_by(announcement_id=meeting_id).all()
    beneficiaries = MeetingBeneficiary.query.filter_by(announcement_id=meeting_id).all()
    attendances = MeetingAttendanceDetail.query.filter_by(announcement_id=meeting_id).all()
    
    return render_template('meeting_detail.html', 
                          meeting=meeting, 
                          agenda_items=agenda_items, 
                          beneficiaries=beneficiaries, 
                          attendances=attendances)


@app.route('/meetings/add', methods=['GET', 'POST'])
@login_required
@role_required(['SECRETAIRE', 'PRESIDENT'])
def add_meeting():
    """Ajouter un nouveau PV de réunion"""
    form = MeetingAttendanceForm()
    set_member_choices(form)
    
    # UTILISER LA VALIDATION DU FORMULAIRE
    if form.validate_on_submit():
        try:
            # Créer la réunion avec les données validées
            meeting = Announcement(
                title=form.meeting_title.data,
                event_date=form.meeting_date.data,
                announcement_type='REUNION',
                is_active=True,
                created_at=datetime.utcnow(),
                created_by=current_user.id,
                content=form.content.data
            )
            db.session.add(meeting)
            db.session.flush()
            
            # Ajouter les points d'ordre du jour (validé)
            if form.agenda_items.data:
                agenda_items = form.agenda_items.data.split('\n')
                for item in agenda_items:
                    if item.strip():
                        agenda = AgendaItem(
                            announcement_id=meeting.id,
                            title=item.strip(),
                            is_completed=False
                        )
                        db.session.add(agenda)
            
            # Ajouter les bénéficiaires (validé)
            if form.beneficiary_member_id.data and form.beneficiary_member_id.data > 0:
                beneficiary = MeetingBeneficiary(
                    announcement_id=meeting.id,
                    member_id=form.beneficiary_member_id.data,
                    benefit_type='TONTINE',
                    amount=Decimal(str(form.benefit_amount.data)) if form.benefit_amount.data else Decimal('0')
                )
                db.session.add(beneficiary)
            
            db.session.commit()
            
            log_activity(current_user.id, current_user.role, f"Création PV réunion: {meeting.title}", request.remote_addr)
            flash('Procès-verbal enregistré avec succès !', 'success')
            return redirect(url_for('meeting_detail', meeting_id=meeting.id))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Erreur : {str(e)}', 'danger')
    else:
        # Afficher les erreurs de validation
        for field, errors in form.errors.items():
            for error in errors:
                field_label = getattr(form, field).label.text if hasattr(form, field) else field
                flash(f'Erreur dans le champ "{field_label}": {error}', 'danger')
    
    return render_template('meeting_form.html', title="Rédiger un PV", form=form)


@app.route('/meetings/<int:meeting_id>/edit', methods=['GET', 'POST'])
@login_required
@role_required(['SECRETAIRE', 'PRESIDENT'])
def edit_meeting(meeting_id):
    """Modifier un PV de réunion"""
    meeting = Announcement.query.get_or_404(meeting_id)
    form = MeetingAttendanceForm(obj=meeting)
    set_member_choices(form)
    
    # UTILISER LA VALIDATION DU FORMULAIRE
    if form.validate_on_submit():
        try:
            meeting.title = form.meeting_title.data
            meeting.event_date = form.meeting_date.data
            meeting.content = form.content.data
            
            # Mettre à jour l'ordre du jour
            AgendaItem.query.filter_by(announcement_id=meeting_id).delete()
            if form.agenda_items.data:
                agenda_items = form.agenda_items.data.split('\n')
                for item in agenda_items:
                    if item.strip():
                        agenda = AgendaItem(
                            announcement_id=meeting.id,
                            title=item.strip(),
                            is_completed=False
                        )
                        db.session.add(agenda)
            
            db.session.commit()
            
            log_activity(current_user.id, current_user.role, f"Modification PV réunion #{meeting_id}", request.remote_addr)
            flash('Procès-verbal modifié avec succès !', 'success')
            return redirect(url_for('meeting_detail', meeting_id=meeting.id))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Erreur : {str(e)}', 'danger')
    else:
        # Afficher les erreurs de validation
        for field, errors in form.errors.items():
            for error in errors:
                field_label = getattr(form, field).label.text if hasattr(form, field) else field
                flash(f'Erreur dans le champ "{field_label}": {error}', 'danger')
    
    # Pré-remplir le formulaire pour la méthode GET
    if request.method == 'GET':
        form.meeting_title.data = meeting.title
        form.meeting_date.data = meeting.event_date
        form.content.data = meeting.content
        agenda_list = '\n'.join([item.title for item in meeting.agenda_items])
        form.agenda_items.data = agenda_list
    
    return render_template('meeting_form.html', title="Modifier le PV", form=form, meeting=meeting)


@app.route('/meetings/<int:meeting_id>/attendance', methods=['GET', 'POST'])
@login_required
@role_required(['SECRETAIRE', 'PRESIDENT'])
def manage_attendance(meeting_id):
    """Gérer les présences à une réunion"""
    meeting = Announcement.query.get_or_404(meeting_id)
    members = Member.query.filter_by(is_active=True, status='ACTIF').order_by(Member.last_name).all()
    
    if request.method == 'POST':
        try:
            # Supprimer les anciennes présences
            MeetingAttendanceDetail.query.filter_by(announcement_id=meeting_id).delete()
            
            # Enregistrer les nouvelles présences avec validation
            for member in members:
                status = request.form.get(f'attendance_{member.id}', 'ABSENT')
                # Validation du statut
                if status not in ['PRESENT', 'ABSENT', 'RETARD', 'EXCUSE']:
                    status = 'ABSENT'
                
                attendance = MeetingAttendanceDetail(
                    announcement_id=meeting_id,
                    member_id=member.id,
                    attendance_status=status
                )
                db.session.add(attendance)
            
            db.session.commit()
            flash('Présences enregistrées avec succès !', 'success')
            return redirect(url_for('meeting_detail', meeting_id=meeting_id))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Erreur lors de l\'enregistrement des présences : {str(e)}', 'danger')
    
    # Récupérer les présences existantes
    existing_attendances = {
        a.member_id: a.attendance_status 
        for a in MeetingAttendanceDetail.query.filter_by(announcement_id=meeting_id).all()
    }
    
    return render_template('meeting_attendance.html', 
                           meeting=meeting, 
                           members=members, 
                           attendances=existing_attendances)


@app.route('/meetings/<int:meeting_id>/delete', methods=['POST'])
@login_required
@role_required(['SECRETAIRE', 'PRESIDENT'])
def delete_meeting(meeting_id):
    """Supprimer une réunion"""
    meeting = Announcement.query.get_or_404(meeting_id)
    
    # Vérifier le token CSRF
    csrf_token = request.form.get('csrf_token')
    if not csrf_token:
        flash('Token CSRF manquant.', 'danger')
        return redirect(url_for('meetings'))
    
    try:
        db.session.delete(meeting)
        db.session.commit()
        
        log_activity(current_user.id, current_user.role, f"Suppression réunion #{meeting_id}", request.remote_addr)
        flash('La réunion a été supprimée avec succès.', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Erreur lors de la suppression : {str(e)}', 'danger')
    
    return redirect(url_for('meetings'))
# ============================================================
# GESTION DES ANNONCES
# ============================================================

@app.route('/annonces')
@login_required
def annonces():
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 10, type=int)
    pagination = Announcement.query.filter(Announcement.announcement_type.in_(['INFO', 'URGENT', 'RAPPEL'])).order_by(Announcement.created_at.desc()).paginate(page=page, per_page=per_page, error_out=False)
    pagination_data = get_pagination_data(pagination)
    return render_template('annonces.html', annonces=pagination_data['items'], pagination=pagination_data)

@app.route('/annonces/add', methods=['GET', 'POST'])
@login_required
@role_required(['SECRETAIRE', 'PRESIDENT', 'COMMUNICATION'])
def add_annonce():
    form = AnnouncementForm()
    
    if form.validate_on_submit():
        try:
            annonce = Announcement(
                title=form.title.data,
                content=form.content.data,
                announcement_type=form.announcement_type.data,
                event_date=form.event_date.data,
                created_by=current_user.id,
                created_at=datetime.now(),
                is_active=True
            )
            
            db.session.add(annonce)
            db.session.commit()
            
            log_activity(current_user.id, current_user.role, f"Création: {annonce.title}", request.remote_addr)
            flash("L'annonce a été publiée avec succès !", "success")
            return redirect(url_for('annonces'))
            
        except Exception as e:
            db.session.rollback()
            flash(f"Erreur technique : {str(e)}", "danger")
    
    return render_template('annonce_form.html', form=form, title="Publier une annonce")
  


@app.route('/annonces/<int:annonce_id>/edit', methods=['GET', 'POST'])
@login_required
@role_required(['SECRETAIRE', 'PRESIDENT', 'COMMUNICATION'])
def edit_annonce(annonce_id):
    """Modifier une annonce existante"""
    annonce = Announcement.query.get_or_404(annonce_id)
    form = AnnouncementForm(obj=annonce)
    
    if form.validate_on_submit():
        try:
            annonce.title = form.title.data
            annonce.content = form.content.data
            annonce.announcement_type = form.announcement_type.data
            annonce.event_date = form.event_date.data if form.event_date.data else None
            
            db.session.commit()
            
            log_activity(current_user.id, current_user.role, f"Modification annonce #{annonce_id}", request.remote_addr)
            flash("L'annonce a été modifiée avec succès !", "success")
            return redirect(url_for('annonces'))
            
        except Exception as e:
            db.session.rollback()
            flash(f"Erreur technique : {str(e)}", "danger")
    
    return render_template('annonce_form.html', form=form, annonce=annonce, title="Modifier l'annonce")


@app.route('/annonces/<int:annonce_id>/delete', methods=['POST'])
@login_required
@role_required(['SECRETAIRE', 'PRESIDENT'])
def delete_annonce(annonce_id):
    """Supprimer une annonce"""
    annonce = Announcement.query.get_or_404(annonce_id)
    
    try:
        db.session.delete(annonce)
        db.session.commit()
        
        log_activity(current_user.id, current_user.role, f"Suppression annonce #{annonce_id}", request.remote_addr)
        flash("L'annonce a été supprimée avec succès.", "success")
        
    except Exception as e:
        db.session.rollback()
        flash(f"Erreur lors de la suppression : {str(e)}", "danger")
    
    return redirect(url_for('annonces'))
# ============================================================
# CYCLES DE TONTINE
# ============================================================

@app.route('/tontine-cycles')
@login_required
@role_required(['SECRETAIRE', 'PRESIDENT', 'TRESORIER'])
def tontine_cycles():
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 10, type=int)
    pagination = TontineCycleDetail.query.order_by(TontineCycleDetail.start_date.desc()).paginate(page=page, per_page=per_page, error_out=False)
    pagination_data = get_pagination_data(pagination)
    
    active_cycle = TontineCycleDetail.query.filter_by(status='EN_COURS').first()
    next_beneficiary = active_cycle.get_next_beneficiary() if active_cycle else None
    
    members_list = Member.query.filter_by(is_active=True, status='ACTIF').order_by(Member.last_name, Member.first_name).all()
    
    return render_template('tontine_cycles.html', 
                           cycles=pagination_data['items'], 
                           pagination=pagination_data,
                           active_cycle=active_cycle,
                           next_beneficiary=next_beneficiary,
                           members_list=members_list)

@app.route('/tontine-cycles/add', methods=['GET', 'POST'])
@login_required
@role_required(['PRESIDENT', 'SECRETAIRE'])
def add_tontine_cycle():
    if request.method == 'POST':
        try:
            amount_per_member = Decimal(request.form.get('amount_per_member', '0'))
        except (ValueError, TypeError):
            flash('Le montant saisi est invalide.', 'danger')
            return redirect(url_for('tontine_cycles'))

        try:
            members_ids = [int(m_id) for m_id in request.form.getlist('members')]
        except ValueError:
            flash('Identifiants de membres invalides.', 'danger')
            return redirect(url_for('tontine_cycles'))

        random_order = request.form.get('random_order') == 'on'
        total_members = len(members_ids)
        
        if total_members == 0:
            flash('Sélectionnez au moins un membre.', 'danger')
            return redirect(url_for('tontine_cycles'))
            
        active_cycle = TontineCycleDetail.query.filter_by(status='EN_COURS').first()
        if active_cycle:
            flash("Un cycle est déjà en cours.", 'danger')
            return redirect(url_for('tontine_cycles'))
        
        last_cycle = TontineCycleDetail.query.order_by(TontineCycleDetail.cycle_number.desc()).first()
        next_cycle_number = (last_cycle.cycle_number + 1) if last_cycle else 1
        
        total_amount = amount_per_member * total_members
        duration_days = (total_members - 1) * 14
        end_date = date.today() + timedelta(days=duration_days)
        
        try:
            cycle = TontineCycleDetail(
                cycle_number=next_cycle_number,
                group_type=str(total_members),
                amount_per_member=amount_per_member,
                total_amount=total_amount,
                start_date=date.today(),
                end_date=end_date,
                status='EN_COURS',
                created_by=current_user.id
            )
            db.session.add(cycle)
            db.session.flush()
            
            db_members = Member.query.filter(Member.id.in_(members_ids)).all()
            members_map = {m.id: m for m in db_members}
            valid_members_ids = [m_id for m_id in members_ids if m_id in members_map]
            
            if random_order:
                ordered_members = valid_members_ids.copy()
                random.shuffle(ordered_members)
                flash("Ordre aléatoire des bénéficiaires.", 'info')
            else:
                sorted_db_members = sorted(
                    [members_map[m_id] for m_id in valid_members_ids],
                    key=lambda x: (x.position_in_group or 999, x.last_name.lower())
                )
                ordered_members = [m.id for m in sorted_db_members]
            
            for position, member_id in enumerate(ordered_members, 1):
                member = members_map.get(member_id)
                if member:
                    member.group_type = str(total_members)
                    member.position_in_group = position
                    member.chosen_tontine_amount = amount_per_member
                    member.has_received_benefit = False
            
            generate_cycle_contributions(cycle.id, ordered_members, amount_per_member)
            
            db.session.commit()
            
            log_activity(current_user.id, current_user.role, f"Création cycle #{cycle.cycle_number}", request.remote_addr)
            
            flash(f'Cycle #{cycle.cycle_number} créé avec {total_members} membres !', 'success')
            return redirect(url_for('tontine_cycle_detail', cycle_id=cycle.id))

        except Exception as e:
            db.session.rollback()
            flash(f"Erreur : {str(e)}", 'danger')
            return redirect(url_for('tontine_cycles'))
    
    members_list = Member.query.filter_by(is_active=True, status='ACTIF').order_by(Member.last_name, Member.first_name).all()
    return render_template('tontine_cycle_form.html', members_list=members_list)

@app.route('/tontine-cycles/<int:cycle_id>')
@login_required
def tontine_cycle_detail(cycle_id):
    cycle = TontineCycleDetail.query.get_or_404(cycle_id)
    beneficiaries = CycleBeneficiary.query.filter_by(cycle_id=cycle_id).order_by(CycleBeneficiary.position).all()
    next_beneficiary = cycle.get_next_beneficiary()
    return render_template('tontine_cycle_detail.html', 
                          cycle=cycle, 
                          beneficiaries=beneficiaries, 
                          next_beneficiary=next_beneficiary)

# ============================================================
# AJOUTEZ CETTE ROUTE MANQUANTE
# ============================================================

@app.route('/tontine-cycles/<int:cycle_id>/register-benefit', methods=['GET', 'POST'])
@login_required
@role_required(['PRESIDENT', 'TRESORIER', 'SECRETAIRE'])
def register_benefit(cycle_id):
    """Enregistrer le bénéfice d'un membre pour un cycle de tontine"""
    cycle = TontineCycleDetail.query.get_or_404(cycle_id)
    next_beneficiary = cycle.get_next_beneficiary()
    
    if not next_beneficiary:
        flash('Tous les membres ont déjà bénéficié de ce cycle !', 'warning')
        return redirect(url_for('tontine_cycle_detail', cycle_id=cycle_id))
    
    form = TontineBenefitForm()
    
    # Remplir les choix du formulaire
    form.member_id.choices = [(next_beneficiary.id, next_beneficiary.full_name)]
    form.cycle_id.choices = [(cycle.id, f"Cycle #{cycle.cycle_number}")]
    
    if request.method == 'POST' or form.validate_on_submit():
        try:
            # Calculer le montant net après déductions
            sanctions_total = Sanction.query.filter_by(
                member_id=next_beneficiary.id, 
                status='PENDING'
            ).with_entities(db.func.sum(Sanction.amount)).scalar() or 0
            
            net_amount = cycle.total_amount - Decimal(str(sanctions_total)) - next_beneficiary.debit_balance
            net_amount = max(Decimal('0.00'), net_amount)
            sanctions_deducted = cycle.total_amount - net_amount
            
            # 1. Enregistrer le bénéfice
            beneficiary = CycleBeneficiary(
                cycle_id=cycle.id,
                member_id=next_beneficiary.id,
                position=next_beneficiary.position_in_group or 0,
                gross_amount=cycle.total_amount,
                net_amount=net_amount,
                sanctions_deducted=sanctions_deducted,
                benefit_date=date.today(),
                payment_status='PAYE',
                payment_mode=request.form.get('payment_mode', 'ESPECE')
            )
            db.session.add(beneficiary)
            
            # 2. Mettre à jour le membre
            next_beneficiary.has_received_benefit = True
            next_beneficiary.amount_to_receive = net_amount
            
            # 3. Enregistrer la transaction financière
            transaction = Transaction(
                member_id=next_beneficiary.id,
                type='BENEFICE_TONTINE',
                amount=net_amount,
                description=f"Bénéfice tontine - Cycle #{cycle.cycle_number} - Position {next_beneficiary.position_in_group}",
                date=date.today(),
                created_by=current_user.id,
                payment_mode=request.form.get('payment_mode', 'ESPECE')
            )
            db.session.add(transaction)
            
            # 4. Vérifier la fin du cycle
            if cycle.beneficiaries_count + 1 >= cycle.total_members:
                cycle.status = 'TERMINE'
                cycle.end_date = date.today()
            
            db.session.commit()
            
            log_activity(current_user.id, current_user.role, 
                        f"Bénéfice tontine enregistré pour {next_beneficiary.full_name} - Cycle #{cycle.cycle_number}", 
                        request.remote_addr)
            
            flash(f"Bénéfice de {net_amount:,.0f} FCFA enregistré pour {next_beneficiary.full_name}.", 'success')
            return redirect(url_for('tontine_cycle_detail', cycle_id=cycle.id))
            
        except Exception as e:
            db.session.rollback()
            flash(f"Erreur lors de l'enregistrement : {str(e)}", 'danger')
    
    return render_template('tontine_benefit_form.html', 
                          form=form, 
                          cycle=cycle, 
                          member=next_beneficiary,
                          sanctions_total=Sanction.query.filter_by(member_id=next_beneficiary.id, status='PENDING').with_entities(db.func.sum(Sanction.amount)).scalar() or 0)


@app.route('/tontine-cycles/<int:cycle_id>/delete', methods=['POST'])
@login_required
@president_required
def delete_tontine_cycle(cycle_id):
    """Supprimer un cycle de tontine"""
    cycle = TontineCycleDetail.query.get_or_404(cycle_id)
    
    if cycle.status == 'TERMINE':
        flash('Impossible de supprimer un cycle terminé.', 'danger')
        return redirect(url_for('tontine_cycles'))
    
    db.session.delete(cycle)
    db.session.commit()
    
    log_activity(current_user.id, current_user.role, f"Suppression cycle tontine #{cycle_id}", request.remote_addr)
    flash('Cycle supprimé avec succès.', 'success')
    return redirect(url_for('tontine_cycles'))

# ============================================================
# RAPPORTS
# ============================================================

@app.route('/reports', methods=['GET', 'POST'])
@login_required
@role_required(['SECRETAIRE', 'PRESIDENT', 'TRESORIER'])
def reports():
    form = ReportForm()
    
    if request.method == 'POST':
        report_type = request.form.get('report_type')
        start_date_str = request.form.get('start_date')
        end_date_str = request.form.get('end_date')
        export_format = request.form.get('format')

        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        except (ValueError, TypeError):
            flash("Dates invalides.", "danger")
            return redirect(url_for('reports'))

        try:
            transactions = Transaction.query.filter(
                Transaction.date >= start_date,
                Transaction.date <= end_date
            ).order_by(Transaction.date.desc()).all()

            if export_format == 'EXCEL':
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Rapport"
                
                ws['A1'] = f"RAPPORT - {report_type.upper()}"
                ws.append([f"Période : {start_date.strftime('%d/%m/%Y')} au {end_date.strftime('%d/%m/%Y')}"])
                ws.append([])
                
                headers = ["Date", "Membre", "Type", "Montant (FCFA)"]
                ws.append(headers)

                total_amount = 0
                for t in transactions:
                    total_amount += t.amount or 0
                    ws.append([
                        t.date.strftime('%d/%m/%Y') if t.date else 'N/A',
                        f"{t.member.first_name} {t.member.last_name}" if t.member else "Inconnu",
                        t.type,
                        t.amount
                    ])
                
                ws.append([])
                ws.append(["TOTAL", "", "", total_amount])

                output = io.BytesIO()
                wb.save(output)
                file_data = output.getvalue()
                output.close()

                return send_file(
                    io.BytesIO(file_data),
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True,
                    download_name=f'rapport_{report_type}_{start_date_str}.xlsx'
                )

            elif export_format == 'PDF':
                flash("Format PDF non disponible pour le moment.", "warning")
                return redirect(url_for('reports'))

        except Exception as e:
            flash(f"Erreur : {str(e)}", "danger")
            return redirect(url_for('reports'))

    total_members = Member.query.filter_by(is_active=True).count()
    total_transactions = Transaction.query.count()
    active_loans = Loan.query.filter_by(status='ACTIF').count()
    pending_loans = Loan.query.filter_by(status='PENDING').count()
    
    total_contributions = db.session.query(db.func.sum(Transaction.amount)).filter(
        Transaction.type.in_(['TONTINE', 'PRESENCE'])
    ).scalar() or 0
    total_sanctions = db.session.query(db.func.sum(Sanction.amount)).filter_by(status='PAYE').scalar() or 0
    total_loans_given = db.session.query(db.func.sum(Loan.amount)).filter(Loan.status.in_(['ACTIF', 'REMBOURSE'])).scalar() or 0
    total_loans_repaid = db.session.query(db.func.sum(Transaction.amount)).filter_by(type='REMBOURSEMENT').scalar() or 0
    
    transactions_by_type = db.session.query(Transaction.type, db.func.sum(Transaction.amount), db.func.count(Transaction.id)).group_by(Transaction.type).all()
    
    top_contributors = db.session.query(
        Member.first_name, Member.last_name, db.func.sum(Transaction.amount).label('total')
    ).join(Transaction, Member.id == Transaction.member_id).filter(Transaction.type.in_(['TONTINE', 'PRESENCE'])).group_by(Member.id).order_by(db.text('total DESC')).limit(10).all()
     
    recent_transactions = Transaction.query.order_by(Transaction.date.desc()).limit(10).all()

    return render_template(
        'reports.html', 
        form=form, 
        total_members=total_members,
        total_transactions=total_transactions,
        active_loans=active_loans,
        pending_loans=pending_loans,
        total_contributions=total_contributions,
        total_sanctions=total_sanctions,
        total_loans_given=total_loans_given,
        total_loans_repaid=total_loans_repaid,
        transactions_by_type=transactions_by_type,
        top_contributors=top_contributors,
        recent_transactions=recent_transactions,
        now=datetime.now()
    )


    #============================================================
    #aides
    # ============================================================
# GESTION DES AIDES SOCIALES
# ============================================================

# ============================================================
# GESTION DES AIDES SOCIALES
# ============================================================

@app.route('/aides')
@login_required
def aides():
    """Page de gestion des aides sociales"""
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 10, type=int)
    status_filter = request.args.get('status', '')
    
    query = Aide.query
    if not current_user.is_admin():
        query = query.filter_by(member_id=current_user.member_id)
    if status_filter:
        query = query.filter_by(status=status_filter)
    
    pagination = query.order_by(Aide.request_date.desc()).paginate(page=page, per_page=per_page, error_out=False)
    pagination_data = get_pagination_data(pagination)

    form = AideForm()
    approval_form = AideApprovalForm()

    if not current_user.is_admin():
        form.member_id.choices = [(current_user.member_id, current_user.member.full_name)]
    else:
        form.member_id.choices = [(m.id, m.full_name) for m in Member.query.filter_by(is_active=True).all()]
    
    return render_template(
        'aides.html', 
        aides=pagination_data['items'], 
        pagination=pagination_data,
        status_filter=status_filter,
        form=form,
        approval_form=approval_form
    )


@app.route('/aides/request', methods=['POST'])
@login_required
def request_aide():
    """Soumettre une demande d'aide"""
    form = AideForm()
    
    if not current_user.is_admin():
        form.member_id.choices = [(current_user.member_id, current_user.member.full_name)]
    else:
        form.member_id.choices = [(m.id, m.full_name) for m in Member.query.filter_by(is_active=True).all()]
    
    if form.validate_on_submit():
        # Vérifier la limite de 3 aides
        aide_count = Aide.query.filter_by(member_id=form.member_id.data, status='APPROVED').count()
        if aide_count >= 3:
            flash('Ce membre a déjà atteint la limite de 3 aides.', 'danger')
            return redirect(url_for('aides'))
        
        aide = Aide(
            member_id=form.member_id.data,
            aide_type=form.aide_type.data,
            amount=Decimal(str(form.amount.data)),
            request_date=date.today(),
            status='PENDING',
            is_paid=False,
            description=form.description.data
        )
        db.session.add(aide)
        db.session.commit()
        log_activity(current_user.id, current_user.role, f"Demande d'aide #{aide.id}", request.remote_addr)
        flash('Demande d\'aide soumise avec succès.', 'success')
        return redirect(url_for('aides'))
    
    # Afficher les erreurs du formulaire
    for field, errors in form.errors.items():
        for error in errors:
            flash(f'Erreur dans le champ {field}: {error}', 'danger')
    
    return redirect(url_for('aides'))


@app.route('/aides/<int:aide_id>/approve', methods=['POST'])
@login_required
@role_required(['PRESIDENT', 'TRESORIER'])
def approve_aide(aide_id):
    """Approuver une demande d'aide"""
    aide = Aide.query.get_or_404(aide_id)
    
    if aide.status != 'PENDING':
        flash('Cette aide ne peut pas être approuvée.', 'warning')
        return redirect(url_for('aides'))
    
    aide.status = 'APPROVED'
    aide.approval_date = date.today()
    aide.approved_by = current_user.id
    db.session.commit()
    
    log_activity(current_user.id, current_user.role, f"Approbation aide #{aide_id}", request.remote_addr)
    flash('Aide approuvée avec succès.', 'success')
    return redirect(url_for('aides'))


@app.route('/aides/<int:aide_id>/reject', methods=['POST'])
@login_required
@role_required(['PRESIDENT', 'TRESORIER'])
def reject_aide(aide_id):
    """Rejeter une demande d'aide"""
    aide = Aide.query.get_or_404(aide_id)
    
    if aide.status != 'PENDING':
        flash('Cette aide ne peut pas être rejetée.', 'warning')
        return redirect(url_for('aides'))
    
    aide.status = 'REJECTED'
    db.session.commit()
    
    log_activity(current_user.id, current_user.role, f"Rejet aide #{aide_id}", request.remote_addr)
    flash('Aide rejetée.', 'warning')
    return redirect(url_for('aides'))


@app.route('/aides/<int:aide_id>/pay', methods=['POST'])
@login_required
@role_required(['TRESORIER', 'PRESIDENT'])
def pay_aide(aide_id):
    """Payer une aide approuvée"""
    aide = Aide.query.get_or_404(aide_id)
    
    if aide.status != 'APPROVED':
        flash('Cette aide doit être approuvée avant paiement.', 'danger')
        return redirect(url_for('aides'))
    
    if aide.is_paid:
        flash('Cette aide a déjà été payée.', 'warning')
        return redirect(url_for('aides'))
    
    aide.is_paid = True
    
    transaction = Transaction(
        member_id=aide.member_id,
        type='AIDE',
        amount=aide.amount,
        description=f"Aide {aide.get_type_display()} - {aide.description or 'Sans description'}",
        date=date.today(),
        created_by=current_user.id,
        payment_mode=request.form.get('payment_mode', 'ESPECE')
    )
    db.session.add(transaction)
    db.session.commit()
    
    log_activity(current_user.id, current_user.role, f"Paiement aide #{aide_id}", request.remote_addr)
    flash('Aide payée avec succès.', 'success')
    return redirect(url_for('aides'))


@app.route('/aides/<int:aide_id>/delete', methods=['POST'])
@login_required
@president_required
def delete_aide(aide_id):
    """Supprimer une demande d'aide (uniquement si non approuvée et non payée)"""
    aide = Aide.query.get_or_404(aide_id)
    
    if aide.status == 'APPROVED':
        flash('Impossible de supprimer une aide déjà approuvée.', 'danger')
        return redirect(url_for('aides'))
    
    if aide.is_paid:
        flash('Impossible de supprimer une aide déjà payée.', 'danger')
        return redirect(url_for('aides'))
    
    db.session.delete(aide)
    db.session.commit()
    
    log_activity(current_user.id, current_user.role, f"Suppression aide #{aide_id}", request.remote_addr)
    flash('Demande d\'aide supprimée.', 'success')
    return redirect(url_for('aides'))
    #============================================================

    #===========================================================
#audit logs
# ============================================================
# MODULE : AUDIT LOGS
# ============================================================

@app.route('/audit-logs')
@login_required
@role_required(['PRESIDENT', 'SECRETAIRE'])
def audit_logs():
    """Journal d'audit des actions des utilisateurs"""
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    user_id = request.args.get('user_id', type=int)
    action_filter = request.args.get('action', '')
    
    query = AuditLog.query
    if user_id:
        query = query.filter_by(user_id=user_id)
    if action_filter:
        query = query.filter(AuditLog.action.ilike(f'%{action_filter}%'))
    
    pagination = query.order_by(AuditLog.timestamp.desc()).paginate(page=page, per_page=per_page, error_out=False)
    pagination_data = get_pagination_data(pagination)
    users = User.query.all()
    
    return render_template('audit_logs.html', 
                          logs=pagination_data['items'], 
                          pagination=pagination_data,
                          users=users, 
                          action_filter=action_filter, 
                          selected_user=user_id)
# ============================================================
# PROFIL UTILISATEUR
# ============================================================

@app.route('/profile', methods=['GET', 'POST'])
@login_required
def profile():
    form = ProfileEditForm()
    
    if form.validate_on_submit():
        user = current_user
        user.email = form.email.data
        
        if current_user.member_id:
            member = current_user.member
            member.first_name = form.first_name.data
            member.last_name = form.last_name.data
            member.phone = form.phone.data
            if form.photo.data:
                filename = save_uploaded_file(form.photo.data, app.config['UPLOAD_FOLDER'])
                if filename:
                    member.photo = filename
        
        db.session.commit()
        flash('Profil mis à jour !', 'success')
        return redirect(url_for('profile'))
    
    if current_user.member_id:
        member = current_user.member
        form.first_name.data = member.first_name
        form.last_name.data = member.last_name
        form.phone.data = member.phone
    form.email.data = current_user.email
    
    return render_template('profile.html', form=form)

# ============================================================
# GESTION DES ERREURS
# ============================================================

@app.errorhandler(403)
def forbidden_error(error):
    return render_template('403.html'), 403

@app.errorhandler(404)
def not_found_error(error):
    return render_template('404.html'), 404

@app.errorhandler(500)
def internal_error(error):
    db.session.rollback()
    return render_template('500.html'), 500

# ============================================================
# POINT D'ENTRÉE
# ============================================================

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)